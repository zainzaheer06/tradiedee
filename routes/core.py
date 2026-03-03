"""
Core Routes Blueprint
Handles authentication, main pages, admin functions, tools management, and webhooks
Includes 30 routes organized in the following categories:
- Authentication (6 routes)
- Main Pages (7 routes)
- Admin (9 routes)
- Tools (5 routes)
- API/Webhooks (3 routes)
"""
import os
import json
import math
import asyncio
import logging
from datetime import datetime, timezone
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from livekit import api

from models import db, User, Agent, CallLog, Tool, AgentTool, InboundConfiguration, Campaign, CampaignContact, Workflow, Business, Job, SMSLog, SAUDI_TZ
from utils.decorators import login_required, admin_required, approved_required
from utils.email import send_verification_email, send_approval_notification, verify_token
from utils.helpers import clean_text
from services.recording_service import recording_service
from services.webhook_service import webhook_service
from services.calltradie_integration import extractor as entity_extractor

logger = logging.getLogger(__name__)

# Create blueprint
core_bp = Blueprint('core', __name__)


# ==================== AUTHENTICATION ROUTES ====================

@core_bp.route('/')
def index():
    """Landing page - redirect to dashboard if logged in, otherwise to login"""
    if 'user_id' in session:
        return redirect(url_for('core.dashboard'))
    return redirect(url_for('core.login'))


@core_bp.route('/homepage')
def homepage():
    """Public homepage"""
    return render_template('index.html')


# Route for modules page
@core_bp.route('/modules')
def modules_page():
    return render_template('modules/modules.html')    


@core_bp.route('/signup', methods=['GET', 'POST'])
def signup():
    """User signup with email verification"""
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'danger')
            return redirect(url_for('core.signup'))

        if User.query.filter_by(email=email).first():
            flash('Email already exists', 'danger')
            return redirect(url_for('core.signup'))

        hashed_password = generate_password_hash(password)
        new_user = User(
            username=username,
            email=email,
            password=hashed_password,
            is_approved=False,
            is_email_verified=False,
            minutes_balance=0
        )

        db.session.add(new_user)
        db.session.commit()

        # Send verification email
        if send_verification_email(email, username):
            flash('Account created! Please check your email to verify your account.', 'success')
        else:
            flash('Account created, but failed to send verification email. Please contact support.', 'warning')

        return redirect(url_for('core.login'))

    return render_template('auth/signup.html')


@core_bp.route('/login', methods=['GET', 'POST'])
def login():
    """User login (accepts username or email)"""
    if request.method == 'POST':
        login_input = request.form.get('username', '').strip()
        password = request.form.get('password')

        # Try username first, then email
        user = User.query.filter_by(username=login_input).first()
        if not user and '@' in login_input:
            user = User.query.filter_by(email=login_input).first()

        if user and check_password_hash(user.password, password):
            session.permanent = True  # use PERMANENT_SESSION_LIFETIME so cookie is saved
            session['user_id'] = user.id
            session['username'] = user.username
            session['is_admin'] = user.is_admin
            session['minutes_balance'] = user.minutes_balance
            flash(f'Welcome back, {user.username}!', 'success')
            return redirect(url_for('core.dashboard'))
        else:
            flash('Invalid username/email or password', 'danger')

    return render_template('auth/login.html')


@core_bp.route('/verify-email/<token>')
def verify_email(token):
    """Verify user email address using the token"""
    email = verify_token(token)

    if not email:
        flash('The verification link is invalid or has expired.', 'danger')
        return redirect(url_for('core.login'))

    user = User.query.filter_by(email=email).first()

    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('core.login'))

    if user.is_email_verified:
        flash('Email already verified. You can log in now.', 'success')
        return redirect(url_for('core.login'))

    user.is_email_verified = True
    user.email_verified_at = datetime.now(SAUDI_TZ).replace(tzinfo=None)
    db.session.commit()

    flash('Email verified successfully! Please wait for admin approval to access your account.', 'success')
    return redirect(url_for('core.login'))


@core_bp.route('/resend-verification', methods=['POST'])
def resend_verification():
    """Resend verification email"""
    email = request.form.get('email')
    user = User.query.filter_by(email=email).first()

    if not user:
        flash('Email not found.', 'danger')
        return redirect(url_for('core.login'))

    if user.is_email_verified:
        flash('Email is already verified.', 'success')
        return redirect(url_for('core.login'))

    if send_verification_email(user.email, user.username):
        flash('Verification email sent! Please check your inbox.', 'success')
    else:
        flash('Failed to send verification email. Please try again later.', 'danger')

    return redirect(url_for('core.login'))


@core_bp.route('/logout')
def logout():
    """Logout user"""
    session.clear()
    flash('Logged out successfully', 'success')
    return redirect(url_for('core.login'))


@core_bp.route('/pending-approval')
@login_required
def pending_approval():
    """Pending approval page for users waiting for admin approval"""
    return render_template('auth/pending_approval.html')


# ==================== MAIN PAGES ====================

@core_bp.route('/dashboard')
@login_required
@approved_required
def dashboard():
    """Owner dashboard - main landing page after login"""
    user = db.session.get(User, session['user_id'])

    if user.is_admin:
        return redirect(url_for('core.admin_dashboard'))

    today_start = datetime.combine(datetime.now().date(), datetime.min.time())
    today_end = datetime.combine(datetime.now().date(), datetime.max.time())

    business = Business.query.filter_by(user_id=user.id).first()

    # Call stats (today)
    total_calls_today = CallLog.query.filter(
        CallLog.user_id == user.id,
        CallLog.created_at >= today_start,
        CallLog.created_at <= today_end
    ).count()

    if business:
        jobs_today = Job.query.filter(
            Job.business_id == business.id,
            Job.created_at >= today_start,
            Job.created_at <= today_end
        ).count()

        emergency_jobs = Job.query.filter(
            Job.business_id == business.id,
            Job.is_emergency == True,
            Job.created_at >= today_start,
            Job.created_at <= today_end
        ).count()

        missed_calls_saved = max(0, total_calls_today - jobs_today)

        recent_jobs = Job.query.filter_by(business_id=business.id)\
            .order_by(Job.created_at.desc()).limit(5).all()
    else:
        jobs_today = 0
        emergency_jobs = 0
        missed_calls_saved = 0
        recent_jobs = []

    stats = {
        'total_calls_today': total_calls_today,
        'jobs_today': jobs_today,
        'emergencies': emergency_jobs,
        'missed_calls_saved': missed_calls_saved,
    }

    return render_template('user/user_dashboard.html',
        user=user, business=business, stats=stats, recent_jobs=recent_jobs, datetime=datetime)


@core_bp.route('/subscription')
@login_required
@approved_required
def subscription():
    """Subscription management page"""
    user = db.session.get(User, session['user_id'])
    return render_template('user/subscription.html', user=user, datetime=datetime)


@core_bp.route('/web-call')
@login_required
@approved_required
def web_call():
    """Web call interface page"""
    user = db.session.get(User, session['user_id'])
    return render_template('calls/web_call.html', user=user)


@core_bp.route('/custom-tools')
@login_required
@approved_required
def custom_tools():
    """Custom tools page"""
    user = db.session.get(User, session['user_id'])
    return render_template('custom_tools.html', user=user)


def _build_call_logs_query(user_id):
    """Build filtered CallLog query from request args. Shared by call_logs page and export."""
    filters = {
        'status': request.args.get('status', ''),
        'call_type': request.args.get('call_type', ''),
        'interest': request.args.get('interest', ''),
        'agent': request.args.get('agent', '', type=str),
        'phone': request.args.get('phone', '').strip(),
        'date_from': request.args.get('date_from', ''),
        'date_to': request.args.get('date_to', ''),
    }

    query = CallLog.query.filter_by(user_id=user_id)

    if filters['status']:
        query = query.filter(CallLog.status == filters['status'])
    if filters['call_type']:
        query = query.filter(CallLog.call_type == filters['call_type'])
    if filters['agent']:
        query = query.filter(CallLog.agent_id == filters['agent'])
    if filters['phone']:
        query = query.filter(
            db.or_(
                CallLog.to_number.contains(filters['phone']),
                CallLog.from_number.contains(filters['phone'])
            )
        )
    if filters['date_from']:
        try:
            date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d')
            query = query.filter(CallLog.created_at >= date_from)
        except ValueError:
            pass
    if filters['date_to']:
        try:
            date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(CallLog.created_at <= date_to)
        except ValueError:
            pass
    if filters['interest']:
        if filters['interest'] == 'interested':
            query = query.filter(CallLog.sentiment_summary.like('%"interest": "Interested"%'))
        elif filters['interest'] == 'not_interested':
            query = query.filter(CallLog.sentiment_summary.like('%"interest": "Not Interested"%'))
        elif filters['interest'] == 'na':
            query = query.filter(db.or_(
                CallLog.status == 'no_answer',
                CallLog.status == 'failed'
            ))

    return query, filters


@core_bp.route('/call-logs')
@login_required
@approved_required
def call_logs():
    """Outbound call logs with pagination and filters"""
    user = db.session.get(User, session['user_id'])

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    if per_page not in [10, 25, 50, 100]:
        per_page = 25

    query, filters = _build_call_logs_query(user.id)

    pagination = query.order_by(CallLog.created_at.desc())\
        .paginate(page=page, per_page=per_page, error_out=False)

    agents = Agent.query.filter_by(user_id=user.id).order_by(Agent.name).all()

    return render_template('calls/call_logs.html',
                         call_logs=pagination.items,
                         pagination=pagination,
                         per_page=per_page,
                         user=user,
                         agents=agents,
                         filters=filters)


@core_bp.route('/call-logs/export')
@login_required
@approved_required
def export_call_logs():
    """Export filtered call logs as Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    user = db.session.get(User, session['user_id'])
    query, filters = _build_call_logs_query(user.id)
    calls = query.order_by(CallLog.created_at.desc()).limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Call Logs"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="E5E7EB")
    )

    headers = ["Date", "Phone Number", "From Number", "Agent", "Call Type",
               "Duration (sec)", "Minutes Used", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row_idx, call in enumerate(calls, 2):
        agent_name = ""
        if call.agent_id:
            agent = db.session.get(Agent, call.agent_id)
            agent_name = agent.name if agent else ""

        ws.cell(row=row_idx, column=1, value=call.created_at.strftime('%Y-%m-%d %H:%M') if call.created_at else "")
        ws.cell(row=row_idx, column=2, value=call.to_number or "")
        ws.cell(row=row_idx, column=3, value=call.from_number or "")
        ws.cell(row=row_idx, column=4, value=agent_name)
        ws.cell(row=row_idx, column=5, value=call.call_type or "outbound")
        ws.cell(row=row_idx, column=6, value=call.duration_seconds or 0)
        ws.cell(row=row_idx, column=7, value=call.minutes_used or 0)
        ws.cell(row=row_idx, column=8, value=call.status or "")

        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).border = thin_border

    # Auto-width columns
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 18

    # Generate file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"call_logs_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@core_bp.route('/call-log/<int:log_id>')
@login_required
@approved_required
def view_call_log(log_id):
    """View individual call log details"""
    log = CallLog.query.get_or_404(log_id)

    if log.user_id != session['user_id'] and not session.get('is_admin'):
        flash('Access denied', 'danger')
        return redirect(url_for('core.dashboard'))

    return render_template('calls/view_call_log.html', log=log)


@core_bp.route('/analytics')
@login_required
@approved_required
def analytics():
    """Analytics dashboard with comprehensive statistics and visualizations"""
    user = db.session.get(User, session['user_id'])

    # Get date range filter (default: last 30 days)
    from datetime import timedelta
    date_range = request.args.get('range', '30')

    # Calculate date boundaries
    end_date = datetime.now(SAUDI_TZ).replace(tzinfo=None)
    if date_range == '12':
        start_date = end_date - timedelta(hours=12)
        range_label = 'Last 12 Hours'
    elif date_range == '24':
        start_date = end_date - timedelta(hours=24)
        range_label = 'Last 24 Hours'
    elif date_range == '7':
        start_date = end_date - timedelta(days=7)
        range_label = 'Last 7 Days'
    elif date_range == '30':
        start_date = end_date - timedelta(days=30)
        range_label = 'Last 30 Days'
    elif date_range == '90':
        start_date = end_date - timedelta(days=90)
        range_label = 'Last 90 Days'
    else:  # all
        start_date = None
        range_label = 'All Time'

    # Base query for user's calls
    base_query = CallLog.query.filter_by(user_id=user.id)
    if start_date:
        # SQLite stores datetimes as naive, so convert start_date to naive for comparison
        start_date_naive = start_date.replace(tzinfo=None)
        base_query = base_query.filter(CallLog.created_at >= start_date_naive)

    # 1. OVERVIEW STATISTICS
    all_calls = base_query.all()
    total_calls = len(all_calls)

    completed_calls = [c for c in all_calls if c.status == 'completed']
    no_answer_calls = [c for c in all_calls if c.status == 'no_answer']
    failed_calls = [c for c in all_calls if c.status == 'failed']

    success_rate = round((len(completed_calls) / total_calls * 100), 1) if total_calls > 0 else 0

    # Average duration (only completed calls)
    if completed_calls:
        avg_duration = round(sum(c.duration_seconds for c in completed_calls) / len(completed_calls), 1)
    else:
        avg_duration = 0

    # Interest rate (% of completed calls that were interested)
    interested_calls = []
    not_interested_calls = []
    for call in completed_calls:
        if call.sentiment_summary and call.sentiment_summary != '{}':
            try:
                sentiment = json.loads(call.sentiment_summary)
                if sentiment.get('interest') == 'Interested':
                    interested_calls.append(call)
                elif sentiment.get('interest') == 'Not Interested':
                    not_interested_calls.append(call)
            except:
                pass

    interest_rate = round((len(interested_calls) / len(completed_calls) * 100), 1) if completed_calls else 0

    # 2. CALLS TREND DATA (grouped by hour for short ranges, by date for longer ranges)
    calls_by_time = {}

    # Use hourly grouping for 12h and 24h ranges, daily for others
    if date_range in ['12', '24']:
        # Group by hour
        for call in all_calls:
            time_key = call.created_at.strftime('%Y-%m-%d %H:00')
            if time_key not in calls_by_time:
                calls_by_time[time_key] = 0
            calls_by_time[time_key] += 1

        # Sort by time and prepare for chart
        sorted_times = sorted(calls_by_time.keys())
        trend_labels = [datetime.strptime(t, '%Y-%m-%d %H:00').strftime('%b %d %H:%M') for t in sorted_times]
        trend_data = [calls_by_time[t] for t in sorted_times]
    else:
        # Group by day
        for call in all_calls:
            date_key = call.created_at.strftime('%Y-%m-%d')
            if date_key not in calls_by_time:
                calls_by_time[date_key] = 0
            calls_by_time[date_key] += 1

        # Sort by date and prepare for chart
        sorted_dates = sorted(calls_by_time.keys())
        trend_labels = [datetime.strptime(d, '%Y-%m-%d').strftime('%b %d') for d in sorted_dates]
        trend_data = [calls_by_time[d] for d in sorted_dates]

    # 3. STATUS BREAKDOWN
    status_data = {
        'completed': len(completed_calls),
        'no_answer': len(no_answer_calls),
        'failed': len(failed_calls)
    }

    # 4. CALL TYPE SPLIT
    inbound_calls = len([c for c in all_calls if c.call_type == 'inbound'])
    outbound_calls = len([c for c in all_calls if c.call_type == 'outbound'])

    # 5. AGENT PERFORMANCE
    agent_stats = {}
    for call in all_calls:
        if call.agent_id:
            if call.agent_id not in agent_stats:
                agent_stats[call.agent_id] = {
                    'agent': call.agent,
                    'total_calls': 0,
                    'completed': 0,
                    'interested': 0,
                    'total_duration': 0
                }

            agent_stats[call.agent_id]['total_calls'] += 1

            if call.status == 'completed':
                agent_stats[call.agent_id]['completed'] += 1
                agent_stats[call.agent_id]['total_duration'] += call.duration_seconds

                # Check interest
                if call.sentiment_summary and call.sentiment_summary != '{}':
                    try:
                        sentiment = json.loads(call.sentiment_summary)
                        if sentiment.get('interest') == 'Interested':
                            agent_stats[call.agent_id]['interested'] += 1
                    except:
                        pass

    # Calculate percentages for agents
    agent_performance = []
    for agent_id, stats in agent_stats.items():
        success_rate_agent = round((stats['completed'] / stats['total_calls'] * 100), 1) if stats['total_calls'] > 0 else 0
        interest_rate_agent = round((stats['interested'] / stats['completed'] * 100), 1) if stats['completed'] > 0 else 0
        avg_duration_agent = round(stats['total_duration'] / stats['completed'], 1) if stats['completed'] > 0 else 0

        agent_performance.append({
            'agent': stats['agent'],
            'total_calls': stats['total_calls'],
            'success_rate': success_rate_agent,
            'interest_rate': interest_rate_agent,
            'avg_duration': avg_duration_agent
        })

    # Sort by total calls (descending)
    agent_performance.sort(key=lambda x: x['total_calls'], reverse=True)

    # 6. RECENT ACTIVITY (last 10 calls)
    recent_calls = CallLog.query.filter_by(user_id=user.id)\
        .order_by(CallLog.created_at.desc())\
        .limit(10)\
        .all()

    return render_template('analytics/analytics.html',
                         user=user,
                         range_label=range_label,
                         selected_range=date_range,
                         # Overview stats
                         total_calls=total_calls,
                         success_rate=success_rate,
                         avg_duration=avg_duration,
                         interest_rate=interest_rate,
                         # Trend data
                         trend_labels=json.dumps(trend_labels),
                         trend_data=json.dumps(trend_data),
                         # Status breakdown
                         status_data=json.dumps(status_data),
                         # Call types
                         inbound_calls=inbound_calls,
                         outbound_calls=outbound_calls,
                         # Agent performance
                         agent_performance=agent_performance,
                         # Recent activity
                         recent_calls=recent_calls,
                         # Interest breakdown
                         interested_count=len(interested_calls),
                         not_interested_count=len(not_interested_calls))


# ==================== ADMIN ROUTES ====================

@core_bp.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    """Admin dashboard with user management"""
    users = User.query.filter_by(is_admin=False).all()
    pending_users = User.query.filter_by(is_approved=False, is_admin=False).all()
    total_calls = CallLog.query.count()
    total_minutes = db.session.query(db.func.sum(CallLog.minutes_used)).scalar() or 0

    return render_template('admin/admin_dashboard.html',
                         users=users,
                         pending_users=pending_users,
                         total_calls=total_calls,
                         total_minutes=total_minutes)


@core_bp.route('/admin/approve-user/<int:user_id>')
@login_required
@admin_required
def approve_user(user_id):
    """Approve a pending user and grant initial minutes"""
    user = User.query.get_or_404(user_id)
    user.is_approved = True
    user.minutes_balance = 10  # Give 10 free minutes
    user.subscription_plan = 'free'
    user.subscription_start_date = datetime.now(SAUDI_TZ).replace(tzinfo=None)
    db.session.commit()

    # Send approval notification email
    send_approval_notification(user.email, user.username)

    flash(f'User {user.username} approved with 10 free minutes', 'success')
    return redirect(url_for('core.admin_dashboard'))


@core_bp.route('/admin/add-minutes/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def add_minutes(user_id):
    """Add minutes to a user's account"""
    user = User.query.get_or_404(user_id)
    minutes = int(request.form.get('minutes', 0))
    user.minutes_balance += minutes
    db.session.commit()
    flash(f'Added {minutes} minutes to {user.username}', 'success')
    return redirect(url_for('core.admin_dashboard'))


@core_bp.route('/admin/trunk-management')
@login_required
@admin_required
def trunk_management():
    """View and manage SIP trunk configuration for all users"""
    users = User.query.filter_by(is_admin=False).order_by(User.username).all()
    return render_template('admin/admin_trunk_management.html', users=users)


@core_bp.route('/admin/configure-trunk/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def configure_user_trunk(user_id):
    """Configure SIP trunk for a specific user"""
    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        outbound_trunk_id = request.form.get('outbound_trunk_id', '').strip()
        sip_notes = request.form.get('sip_notes', '').strip()

        # Update user trunk configuration
        user.outbound_trunk_id = outbound_trunk_id if outbound_trunk_id else None
        user.sip_configured = bool(outbound_trunk_id)
        user.sip_configured_at = datetime.now(SAUDI_TZ).replace(tzinfo=None) if outbound_trunk_id else None
        user.sip_notes = sip_notes if sip_notes else None

        db.session.commit()

        if outbound_trunk_id:
            flash(f'SIP trunk configured for {user.username}', 'success')
        else:
            flash(f'SIP trunk removed for {user.username}', 'success')

        return redirect(url_for('core.trunk_management'))

    return render_template('admin/admin_configure_trunk.html', user=user)


@core_bp.route('/admin/remove-trunk/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def remove_user_trunk(user_id):
    """Remove SIP trunk configuration from a user"""
    user = User.query.get_or_404(user_id)

    user.outbound_trunk_id = None
    user.sip_configured = False
    user.sip_configured_at = None
    user.sip_notes = None

    db.session.commit()
    flash(f'SIP trunk removed from {user.username}', 'success')
    return redirect(url_for('core.trunk_management'))


@core_bp.route('/admin/inbound-trunk-management')
@login_required
@admin_required
def inbound_trunk_management():
    """View and manage inbound SIP trunk configuration for inbound configurations"""
    # Get all inbound configurations from all users
    inbound_configs = InboundConfiguration.query.order_by(InboundConfiguration.created_at.desc()).all()
    return render_template('admin/admin_inbound_trunk_management.html', configs=inbound_configs)


@core_bp.route('/admin/configure-inbound-trunk/<int:config_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def configure_inbound_trunk(config_id):
    """Configure inbound SIP trunk for a specific inbound configuration"""
    config = InboundConfiguration.query.get_or_404(config_id)

    if request.method == 'POST':
        inbound_trunk_id = request.form.get('inbound_trunk_id', '').strip()

        # Update configuration trunk
        config.trunk_id = inbound_trunk_id if inbound_trunk_id else None
        db.session.commit()

        if inbound_trunk_id:
            flash(f'Inbound trunk "{inbound_trunk_id}" configured for "{config.name}"', 'success')
        else:
            flash(f'Inbound trunk removed from "{config.name}"', 'success')

        return redirect(url_for('core.inbound_trunk_management'))

    return render_template('admin/admin_configure_inbound_trunk.html', config=config)


@core_bp.route('/admin/remove-inbound-trunk/<int:config_id>', methods=['POST'])
@login_required
@admin_required
def remove_inbound_trunk(config_id):
    """Remove inbound SIP trunk from inbound configuration"""
    config = InboundConfiguration.query.get_or_404(config_id)

    config.trunk_id = None
    db.session.commit()

    flash(f'Inbound trunk removed from "{config.name}"', 'success')
    return redirect(url_for('core.inbound_trunk_management'))


# ==================== TOOL MANAGEMENT ROUTES ====================

@core_bp.route('/tools', methods=['GET'])
@login_required
@approved_required
def tools_list():
    """List all tools for the current user"""
    tools = Tool.query.filter_by(user_id=session['user_id']).order_by(Tool.created_at.desc()).all()
    return render_template('tools/list.html', tools=tools)


@core_bp.route('/tools/create', methods=['GET', 'POST'])
@login_required
@approved_required
def create_tool():
    """Create a new tool"""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        tool_type = request.form.get('tool_type', '')

        # Validation
        if not name or not description or not tool_type:
            flash('Name, description, and tool type are required', 'error')
            return render_template('tools/create.html')

        if tool_type not in ['api_call', 'webhook', 'rpc']:
            flash('Invalid tool type', 'error')
            return render_template('tools/create.html')

        # Parse parameters (optional)
        parameters = {}
        parameters_str = request.form.get('parameters', '').strip()
        if parameters_str:
            try:
                parameters = json.loads(parameters_str)
            except json.JSONDecodeError:
                flash('Invalid JSON format for parameters', 'error')
                return render_template('tools/create.html')

        # Build config based on tool type
        config = {}
        if tool_type == 'api_call':
            config = {
                'url': request.form.get('api_url', '').strip(),
                'method': request.form.get('api_method', 'GET'),
                'headers': json.loads(request.form.get('api_headers', '{}')),
                'parameters': parameters
            }
            if not config['url']:
                flash('API URL is required for API call tools', 'error')
                return render_template('tools/create.html')

        elif tool_type == 'webhook':
            config = {
                'url': request.form.get('webhook_url', '').strip(),
                'headers': json.loads(request.form.get('webhook_headers', '{}')),
                'parameters': parameters
            }
            if not config['url']:
                flash('Webhook URL is required for webhook tools', 'error')
                return render_template('tools/create.html')

        elif tool_type == 'rpc':
            config = {
                'method': request.form.get('rpc_method', name),
                'timeout': float(request.form.get('rpc_timeout', 5.0)),
                'parameters': parameters
            }

        # Create tool
        tool = Tool(
            user_id=session['user_id'],
            name=name,
            description=description,
            tool_type=tool_type,
            config=json.dumps(config),
            is_active=True
        )

        db.session.add(tool)
        db.session.commit()

        flash(f'Tool "{name}" created successfully!', 'success')
        return redirect(url_for('core.tools_list'))

    return render_template('tools/create.html')


@core_bp.route('/tools/<int:tool_id>/edit', methods=['GET', 'POST'])
@login_required
@approved_required
def edit_tool(tool_id):
    """Edit an existing tool"""
    tool = Tool.query.filter_by(id=tool_id, user_id=session['user_id']).first_or_404()

    if request.method == 'POST':
        tool.name = request.form.get('name', '').strip()
        tool.description = request.form.get('description', '').strip()
        tool.is_active = request.form.get('is_active') == 'on'

        # Parse parameters (optional)
        parameters = {}
        parameters_str = request.form.get('parameters', '').strip()
        if parameters_str:
            try:
                parameters = json.loads(parameters_str)
            except json.JSONDecodeError:
                flash('Invalid JSON format for parameters', 'error')
                config = json.loads(tool.config)
                return render_template('tools/edit.html', tool=tool, config=config)

        # Update config based on tool type
        config = json.loads(tool.config)
        if tool.tool_type == 'api_call':
            config['url'] = request.form.get('api_url', '').strip()
            config['method'] = request.form.get('api_method', 'GET')
            config['headers'] = json.loads(request.form.get('api_headers', '{}'))
            config['parameters'] = parameters
        elif tool.tool_type == 'webhook':
            config['url'] = request.form.get('webhook_url', '').strip()
            config['headers'] = json.loads(request.form.get('webhook_headers', '{}'))
            config['parameters'] = parameters
        elif tool.tool_type == 'rpc':
            config['method'] = request.form.get('rpc_method', tool.name)
            config['timeout'] = float(request.form.get('rpc_timeout', 5.0))
            config['parameters'] = parameters

        tool.config = json.dumps(config)
        tool.updated_at = datetime.now(SAUDI_TZ).replace(tzinfo=None)

        db.session.commit()
        flash(f'Tool "{tool.name}" updated successfully!', 'success')
        return redirect(url_for('core.tools_list'))

    # Parse config for template
    config = json.loads(tool.config)
    return render_template('tools/edit.html', tool=tool, config=config)


@core_bp.route('/tools/<int:tool_id>/delete', methods=['POST'])
@login_required
@approved_required
def delete_tool(tool_id):
    """Delete a tool"""
    tool = Tool.query.filter_by(id=tool_id, user_id=session['user_id']).first_or_404()

    # Remove all agent associations
    AgentTool.query.filter_by(tool_id=tool_id).delete()

    db.session.delete(tool)
    db.session.commit()

    flash(f'Tool "{tool.name}" deleted successfully!', 'success')
    return redirect(url_for('core.tools_list'))


@core_bp.route('/tools/<int:tool_id>/toggle', methods=['POST'])
@login_required
@approved_required
def toggle_tool(tool_id):
    """Toggle tool active status"""
    tool = Tool.query.filter_by(id=tool_id, user_id=session['user_id']).first_or_404()
    tool.is_active = not tool.is_active
    tool.updated_at = datetime.now(SAUDI_TZ).replace(tzinfo=None)
    db.session.commit()

    status = "activated" if tool.is_active else "deactivated"
    flash(f'Tool "{tool.name}" {status}!', 'success')
    return redirect(url_for('core.tools_list'))


# ==================== API & WEBHOOK ROUTES ====================

@core_bp.route('/api/demo-token', methods=['POST'])
def get_demo_token():
    """Generate a LiveKit token for demo voice agent session"""
    try:
        import secrets

        # Get LiveKit credentials from environment
        livekit_url = os.environ.get('LIVEKIT_URL')
        api_key = os.environ.get('LIVEKIT_API_KEY')
        api_secret = os.environ.get('LIVEKIT_API_SECRET')

        if not all([livekit_url, api_key, api_secret]):
            return jsonify({'error': 'LiveKit credentials not configured'}), 500

        # Generate unique room name for demo
        room_name = f"demo-{secrets.token_hex(8)}"
        participant_identity = f"demo-user-{secrets.token_hex(4)}"

        # Create token
        token = api.AccessToken(api_key, api_secret) \
            .with_identity(participant_identity) \
            .with_name("Demo User") \
            .with_grants(api.VideoGrants(
                room_join=True,
                room=room_name,
                can_publish=True,
                can_subscribe=True,
            ))

        jwt_token = token.to_jwt()

        logger.info(f"<� Generated demo token for room: {room_name}")

        return jsonify({
            'token': jwt_token,
            'url': livekit_url,
            'room_name': room_name,
            'participant_identity': participant_identity
        })

    except Exception as e:
        logger.error(f"Error generating demo token: {e}")
        return jsonify({'error': str(e)}), 500


@core_bp.route('/api/start-demo-agent', methods=['POST'])
def start_demo_agent():
    """Create agent dispatch for demo room"""
    try:
        data = request.json
        room_name = data.get('room_name')

        if not room_name:
            return jsonify({'error': 'room_name required'}), 400

        # Run async function to create dispatch
        async def create_demo_dispatch():
            lkapi = api.LiveKitAPI()

            # Create agent dispatch for the demo room
            dispatch = await lkapi.agent_dispatch.create_dispatch(
                api.CreateAgentDispatchRequest(
                    agent_name="agent-web",  # Use agent-web.py for web demos
                    room=room_name,
                    metadata="demo_session"
                )
            )

            await lkapi.aclose()
            return dispatch

        # Execute async function
        dispatch = asyncio.run(create_demo_dispatch())

        logger.info(f"> Agent dispatched to demo room: {room_name}")

        return jsonify({
            'success': True,
            'dispatch_id': dispatch.id,
            'room_name': room_name
        })

    except Exception as e:
        logger.error(f"Error starting demo agent: {e}")
        return jsonify({'error': str(e)}), 500


@core_bp.route('/webhook/call-ended', methods=['POST'])
def call_ended_webhook():
    """Webhook endpoint to receive call completion data from LiveKit"""
    try:
        data = request.json
        logger.info(f"=� Webhook received: {data}")

        room_name = data.get('room_name')
        duration = data.get('duration', 0)
        transcription = data.get('transcription', '')
        metadata = data.get('metadata', {})
        message_count = data.get('message_count', 0)

        logger.info(f"Processing: {room_name}, {message_count} messages, {duration}s")

        # Check if this is a campaign call (format: campaign_{id}_contact_{id}_{timestamp})
        is_campaign_call = room_name.startswith('campaign_')

        if is_campaign_call:
            logger.info(f"=� Detected campaign call: {room_name}")
            # Parse campaign and contact IDs
            try:
                parts = room_name.split('_')
                campaign_id = int(parts[1])
                contact_id = int(parts[3])

                # Update campaign contact
                contact = db.session.get(CampaignContact, contact_id)
                if contact:
                    contact.duration = duration
                    contact.transcription = transcription

                    # Determine call status - only mark completed if USER spoke
                    messages = metadata.get('messages', [])
                    user_messages = [msg for msg in messages if msg.get('role') == 'user']
                    sip_info = metadata.get('sip_info', {})
                    sip_call_status = sip_info.get('call_status', 'unknown')

                    logger.info(f"= SIP Status: {sip_call_status}, Total messages: {len(messages)}, User messages: {len(user_messages)}, Duration: {duration}s")

                    # Only mark as completed if USER actually spoke
                    if len(user_messages) > 0:
                        contact.status = 'completed'
                        logger.info(f" Campaign call completed ({len(messages)} messages, {len(user_messages)} from user)")

                        # Analyze interest with agent context
                        logger.info(f"=Analyzing customer interest...")
                        # Get agent prompt for context
                        campaign = db.session.get(Campaign, campaign_id)
                        agent_prompt = None
                        if campaign and campaign.agent_id:
                            agent = db.session.get(Agent, campaign.agent_id)
                            if agent:
                                agent_prompt = agent.prompt

                        sentiment_data = analyze_sentiment(messages, agent_prompt=agent_prompt)
                        contact.interest_level = sentiment_data.get('interest', 'Not Interested')
                        logger.info(f" Interest: {contact.interest_level}")
                    elif len(messages) > 0 and len(user_messages) == 0:
                        # Agent spoke (greeting) but user didn't answer
                        contact.status = 'no_answer'
                        logger.info(f"= No answer (agent greeting sent but no user response)")
                    elif sip_call_status in ['dialing', 'ringing']:
                        contact.status = 'no_answer'
                        logger.info(f"= No answer (SIP: {sip_call_status})")
                    else:
                        contact.status = 'failed'
                        logger.info(f"L Call failed (SIP: {sip_call_status})")

                    # Clear transcription for no-answer and failed calls (greeting-only)
                    if contact.status in ['no_answer', 'failed']:
                        contact.transcription = ''
                        contact.transcription_data = '{}'  # Clear JSON metadata too
                        contact.interest_level = None  # Clear interest analysis
                        transcription = ''  # Update local variable for response
                        message_count = 0  # Update count for response
                        logger.info(f"   Cleared transcription, metadata, and interest analysis (no user interaction)")

                    db.session.commit()
                    logger.info(f" Campaign contact updated: {contact_id}")

                    # ALSO CREATE CALL LOG ENTRY for campaign calls
                    campaign = db.session.get(Campaign, campaign_id)
                    if campaign:
                        # Process recording if available (only for completed calls where user spoke)
                        recording_url = None
                        recording_id = metadata.get('recording_id')

                        if recording_id and len(user_messages) > 0:
                            logger.info(f"<� Processing recording {recording_id} for campaign call...")
                            try:
                                # Process recording: download from LiveKit, upload to OSS
                                recording_url = asyncio.run(
                                    recording_service.process_recording(recording_id, room_name)
                                )
                                if recording_url:
                                    logger.info(f" Recording saved to OSS: {recording_url[:100]}...")
                                else:
                                    logger.warning("� Failed to process recording")
                            except Exception as e:
                                logger.error(f"L Error processing recording: {e}")

                        # Only charge if USER spoke (not just agent greeting)
                        # Clear transcription/metadata if no user interaction
                        call_transcription = transcription if len(user_messages) > 0 else ''
                        call_metadata = json.dumps(metadata, ensure_ascii=False) if len(user_messages) > 0 else '{}'


                        # Get user's configured phone number (consistent with agents.py)
                        campaign_user = db.session.get(User, campaign.user_id)
                        user_from_number = (
                            campaign_user.outbound_phone_number 
                            if campaign_user and campaign_user.outbound_phone_number 
                            else sip_info.get('trunk_phone_number', 'campaign')
                        )
    

                        call_log = CallLog(
                            user_id=campaign.user_id,
                            agent_id=campaign.agent_id,
                            from_number=user_from_number,
                            to_number=sip_info.get('phone_number', contact.phone_number),
                            room_name=room_name,
                            duration_seconds=duration,
                            transcription=call_transcription,
                            transcription_data=call_metadata,
                            call_type='outbound',
                            status='completed' if len(user_messages) > 0 else 'no_answer',
                            minutes_used=math.ceil(duration / 60) if len(user_messages) > 0 else 0,
                            recording_url=recording_url,
                            recording_id=recording_id,
                            created_at=datetime.now(SAUDI_TZ).replace(tzinfo=None)
                        )

                        # Add sentiment analysis for campaign calls too with agent context (only if user spoke)
                        if messages and len(user_messages) > 0:
                            # Get agent prompt for context
                            agent_prompt = None
                            if campaign.agent_id:
                                agent = db.session.get(Agent, campaign.agent_id)
                                if agent:
                                    agent_prompt = agent.prompt

                            sentiment_data = analyze_sentiment(messages, agent_prompt=agent_prompt)
                            call_log.sentiment_summary = json.dumps(sentiment_data, ensure_ascii=False)

                        db.session.add(call_log)
                        db.session.commit()
                        logger.info(f" Campaign call log created: {call_log.id}")

                    # Trigger n8n workflow for campaign call (NON-BLOCKING)
                    if campaign.agent_id and contact.status == 'completed':
                        agent = db.session.get(Agent, campaign.agent_id)

                        if agent and agent.workflow_id:
                            workflow = db.session.get(Workflow, agent.workflow_id)

                            if workflow and workflow.is_active:
                                # Prepare comprehensive data for n8n
                                n8n_data = {
                                    'event_type': 'call_completed',
                                    'call_id': call_log.id,
                                    'room_name': room_name,
                                    'agent': {
                                        'id': agent.id,
                                        'name': agent.name
                                    },
                                    'call_details': {
                                        'from_number': sip_info.get('trunk_phone_number', 'campaign'),
                                        'to_number': sip_info.get('phone_number', contact.phone_number),
                                        'duration_seconds': duration,
                                        'duration_minutes': math.ceil(duration / 60) if len(user_messages) > 0 else 0,
                                        'status': contact.status,
                                        'call_type': 'outbound',
                                        'timestamp': call_log.created_at.isoformat()
                                    },
                                    'conversation': {
                                        'transcription': call_transcription,
                                        'message_count': message_count,
                                        'recording_url': recording_url
                                    },
                                    'campaign': {
                                        'campaign_id': campaign_id,
                                        'contact_id': contact_id,
                                        'contact_name': contact.name
                                    }
                                }

                                # Add sentiment/interest analysis
                                if contact.interest_level:
                                    n8n_data['analysis'] = {
                                        'interest_level': contact.interest_level,
                                        'confidence': 'High',
                                        'key_indicators': {}
                                    }

                                # Update workflow stats
                                workflow.total_calls += 1
                                db.session.commit()

                                # Trigger POST-call webhook if enabled (NON-BLOCKING - returns immediately)
                                if workflow.post_call_enabled:
                                    logger.info(f"Triggering workflow for campaign call: {workflow.name}")
                                    webhook_service.trigger_webhook(
                                        workflow_id=workflow.id,
                                        workflow_url=workflow.webhook_url,
                                        api_key=workflow.api_key,
                                        call_data=n8n_data,
                                        call_log_id=call_log.id
                                    )
                                else:
                                    logger.info(f"Post-call webhook DISABLED for campaign workflow: {workflow.name}")

                    return jsonify({
                        'success': True,
                        'message': f'Campaign call processed: {message_count} messages',
                        'contact_id': contact_id
                    }), 200
                else:
                    logger.error(f"L Campaign contact {contact_id} not found")
                    return jsonify({'error': 'Campaign contact not found'}), 404
            except (IndexError, ValueError) as e:
                logger.error(f"L Failed to parse campaign room name: {room_name}, error: {e}")
                return jsonify({'error': 'Invalid campaign room name'}), 400

        # Find call log by room name (for inbound calls)
        call_log = CallLog.query.filter_by(room_name=room_name).first()

        # If call log doesn't exist, create it (inbound call)
        if not call_log:
            logger.info(f"=� Creating new inbound call log for: {room_name}")

            # Extract SIP info
            sip_info = metadata.get('sip_info', {})
            from_number = sip_info.get('phone_number', 'unknown')
            to_number = sip_info.get('trunk_phone_number', 'unknown')

            # Find inbound configuration by phone number called (to_number)
            inbound_config = InboundConfiguration.query.filter_by(phone_number=to_number).first()

            # Fallback: Check for CallTradie Business configuration with matching phone number
            business_config = None
            if not inbound_config:
                business_config = Business.query.filter_by(phone_number=to_number).first()

            # Check if it's a test/demo call
            is_test_call = to_number == 'unknown' or (room_name.startswith('call-') and '-test-' in room_name)

            if is_test_call:
                logger.info(f"= Test call detected ({room_name}), creating test call log")

                # Extract agent_id from room name: call-{agent_id}-test-{timestamp}
                test_agent_id = None
                test_user_id = None
                try:
                    parts = room_name.split('-')
                    if len(parts) >= 2:
                        test_agent_id = int(parts[1])
                        test_agent = db.session.get(Agent, test_agent_id)
                        if test_agent:
                            test_user_id = test_agent.user_id
                            logger.info(f"✅ Test call linked to agent #{test_agent_id} ({test_agent.name}), user_id={test_user_id}")
                except (ValueError, IndexError):
                    pass

                # Fallback: find user by email
                if not test_user_id:
                    test_user = User.query.filter_by(email='example@gmail.com').first()
                    if not test_user:
                        test_user = User.query.first()
                    test_user_id = test_user.id if test_user else 1

                # Create test call log linked to the correct user
                call_log = CallLog(
                    user_id=test_user_id,
                    agent_id=test_agent_id,
                    room_name=room_name,
                    from_number=from_number,
                    to_number='test',
                    call_type='test',
                    status='in_progress',
                    created_at=datetime.now(SAUDI_TZ).replace(tzinfo=None)
                )
                db.session.add(call_log)
                db.session.flush()
                logger.info(f"✅ Created test call log (ID: {call_log.id}) for user ID: {test_user_id}, agent: {test_agent_id}")
            elif inbound_config:
                # Get the linked agent
                agent = inbound_config.agent

                call_log = CallLog(
                    user_id=inbound_config.user_id,
                    agent_id=agent.id,
                    room_name=room_name,
                    from_number=from_number,
                    to_number=to_number,
                    call_type='inbound',
                    status='in_progress',
                    created_at=datetime.now(SAUDI_TZ).replace(tzinfo=None)
                )
                db.session.add(call_log)
                db.session.flush()  # Get the ID
                logger.info(f" Created inbound call log (ID: {call_log.id}) for config: {inbound_config.name}, agent: {agent.name}")
            elif business_config:
                # CallTradie: Create call log for business (no specific agent)
                logger.info(f"✅ Found CallTradie business: {business_config.business_name}")
                call_log = CallLog(
                    user_id=business_config.user_id,
                    agent_id=None,  # CallTradie uses voice system, not agent
                    room_name=room_name,
                    from_number=from_number,
                    to_number=to_number,
                    call_type='inbound',
                    status='in_progress',
                    created_at=datetime.now(SAUDI_TZ).replace(tzinfo=None)
                )
                db.session.add(call_log)
                db.session.flush()  # Get the ID
                logger.info(f" Created inbound call log (ID: {call_log.id}) for CallTradie business: {business_config.business_name}")
            else:
                logger.error(f"L No inbound configuration or business found for phone number: {to_number}")
                return jsonify({'error': f'No inbound configuration for {to_number}'}), 404

        if call_log:
            call_log.duration_seconds = duration
            call_log.transcription = transcription
            call_log.transcription_data = json.dumps(metadata, ensure_ascii=False)  # Save full metadata

            # Determine call status - prioritize USER messages (not just agent greeting)
            messages = metadata.get('messages', [])
            user_messages = [msg for msg in messages if msg.get('role') == 'user']
            sip_info = metadata.get('sip_info', {})
            sip_call_status = sip_info.get('call_status', 'unknown')

            logger.info(f"=� SIP Status: {sip_call_status}, Total messages: {len(messages)}, User messages: {len(user_messages)}, Duration: {duration}s")

            # Save old minutes BEFORE updating (to prevent double-charging on duplicate webhooks)
            old_minutes_used = call_log.minutes_used
            old_status = call_log.status

            # Primary rule: Only charge if USER spoke (actual conversation happened)
            if len(user_messages) > 0:
                # User responded - call was completed and should be charged
                call_log.status = 'completed'
                call_log.minutes_used = math.ceil(duration / 60)
                logger.info(f" Call completed ({len(messages)} messages, {len(user_messages)} from user, duration: {duration}s, minutes: {call_log.minutes_used})")

            # Secondary rule: Agent spoke (greeting) but user didn't respond
            elif len(messages) > 0 and len(user_messages) == 0:
                # Only agent greeting sent, user never answered - no charge
                call_log.status = 'no_answer'
                call_log.minutes_used = 0
                logger.info(f"=� No answer (agent greeting sent but no user response, SIP: {sip_call_status})")

            # Tertiary rule: Check SIP status if no messages at all
            elif sip_call_status in ['dialing', 'ringing']:
                # Call was never answered - no charge
                call_log.status = 'no_answer'
                call_log.minutes_used = 0
                logger.info(f"=� No answer (SIP: {sip_call_status}, no messages)")

            else:
                # No messages and no clear SIP status - technical failure
                call_log.status = 'failed'
                call_log.minutes_used = 0
                logger.info(f"L Call failed (no messages, SIP: {sip_call_status})")

            # Perform interest analysis on conversation with agent context (only if user spoke)
            messages = metadata.get('messages', [])
            user_messages = [msg for msg in messages if msg.get('role') == 'user']
            if messages and len(user_messages) > 0:
                logger.info(f"Analyzing customer interest for {len(messages)} messages...")
                # Get agent prompt for context
                agent_prompt = None
                if call_log.agent_id:
                    agent = db.session.get(Agent, call_log.agent_id)
                    if agent:
                        agent_prompt = agent.prompt

                sentiment_data = analyze_sentiment(messages, agent_prompt=agent_prompt)
                call_log.sentiment_summary = json.dumps(sentiment_data, ensure_ascii=False)
                logger.info(f" Interest analysis complete: {sentiment_data.get('interest', 'Not Interested')}")
            else:
                logger.info("9  No messages to analyze for interest")
                call_log.sentiment_summary = json.dumps({
                    'interest': 'Not Interested',
                    'confidence': 'Low',
                    'key_indicators': {
                        'asked_questions': False,
                        'requested_info': False,
                        'positive_engagement': False,
                        'requested_callback': False
                    }
                }, ensure_ascii=False)

            # Process recording if available (only for completed calls)
            recording_id = metadata.get('recording_id')
            if recording_id and call_log.status == 'completed':
                logger.info(f"<� Processing recording {recording_id} for call...")
                try:
                    # Process recording: download from LiveKit, upload to OSS
                    recording_url = asyncio.run(
                        recording_service.process_recording(recording_id, room_name)
                    )
                    if recording_url:
                        call_log.recording_url = recording_url
                        call_log.recording_id = recording_id
                        logger.info(f" Recording saved to OSS: {recording_url[:100]}...")
                    else:
                        logger.warning("� Failed to process recording")
                except Exception as e:
                    logger.error(f"L Error processing recording: {e}")

            # Deduct minutes from user - ONLY charge the DIFFERENCE (prevents double-charging on duplicate webhooks)
            minutes_difference = call_log.minutes_used - old_minutes_used

            if minutes_difference != 0:
                user = db.session.get(User, call_log.user_id)
                if user:
                    # Adjust user balance by the difference (can be positive or negative)
                    user.minutes_balance -= minutes_difference
                    if user.minutes_balance < 0:
                        user.minutes_balance = 0
                    # Track total minutes used (add the difference)
                    user.minutes_used = (user.minutes_used or 0) + minutes_difference

                    if minutes_difference > 0:
                        logger.info(f"Deducted {minutes_difference} minutes from user {user.username} (was {old_minutes_used}, now {call_log.minutes_used})")
                    else:
                        logger.info(f"Refunded {-minutes_difference} minutes to user {user.username} (was {old_minutes_used}, now {call_log.minutes_used})")
            else:
                logger.info(f"No minutes change (status: {call_log.status}, minutes: {call_log.minutes_used}, old: {old_minutes_used})")

            # Update agent last_used timestamp
            if call_log.agent_id:
                agent = db.session.get(Agent, call_log.agent_id)
                if agent:
                    agent.last_used = datetime.now(SAUDI_TZ).replace(tzinfo=None)

            # Clear transcription for no-answer and failed calls (greeting-only)
            if call_log.status in ['no_answer', 'failed']:
                call_log.transcription = ''
                call_log.transcription_data = '{}'  # Clear JSON metadata too
                call_log.sentiment_summary = '{}'  # Clear interest analysis
                transcription = ''  # Update local variable for response
                message_count = 0  # Update count for response
                logger.info(f"   Cleared transcription, metadata, and interest analysis (no user interaction)")

            db.session.commit()
            logger.info(f" Call log updated: {room_name}")
            # ===== CALLTRADIE JOB CREATION (POST-CALL AI ANALYSIS) =====
            # Analyze completed calls to check if job/booking is needed
            if call_log.status == 'completed' and call_log.call_type in ['inbound', 'test']:
                try:
                    call_type_label = 'test call' if call_log.call_type == 'test' else 'inbound call'
                    logger.info(f"🎯 Analyzing {call_type_label} for job creation...")

                    # Step 1: Use OpenAI to analyze transcript and extract details
                    extracted = entity_extractor.extract_from_transcript_openai_sync(transcription)

                    logger.info(f"📊 AI Analysis: name={extracted.get('customer_name')}, phone={extracted.get('customer_phone')}, service={extracted.get('service_type')}, booking_needed={extracted.get('booking_needed')}, urgency={extracted.get('urgency')}")

                    # Step 2: Decide if job creation is needed
                    booking_needed = extracted.get('booking_needed', False)
                    has_service = extracted.get('service_type') is not None
                    has_issue = extracted.get('issue_description') is not None
                    is_emergency = extracted.get('urgency') == 'emergency'

                    should_create_job = booking_needed or has_service or has_issue or is_emergency

                    if not should_create_job:
                        logger.info(f"ℹ️  No job needed for {call_type_label} - customer didn't request a service or booking")
                    else:
                        reason = []
                        if booking_needed: reason.append("booking requested")
                        if has_service: reason.append(f"service: {extracted.get('service_type')}")
                        if is_emergency: reason.append("EMERGENCY")
                        if has_issue: reason.append("issue described")
                        logger.info(f"✅ Job creation triggered: {', '.join(reason)}")

                        # Step 3: Find business (for real calls) or use None (for test calls)
                        business_id = None
                        if call_log.user_id:
                            business = Business.query.filter_by(user_id=call_log.user_id).first()
                            if business:
                                business_id = business.id
                                logger.info(f"✅ Linked to business: {business.business_name}")

                        # Step 4: Create job with extracted data
                        job = Job(
                            business_id=business_id,
                            original_call_id=call_log.id,
                            customer_name=extracted.get('customer_name'),
                            customer_phone=extracted.get('customer_phone'),
                            customer_address=extracted.get('customer_address'),
                            job_type=extracted.get('service_type', 'general'),
                            description=extracted.get('issue_description', ''),
                            urgency=extracted.get('urgency', 'normal'),
                            is_emergency=is_emergency,
                            call_transcript=transcription,
                            call_summary=f"Call from {call_log.from_number} ({duration}s) [{call_type_label.upper()}]",
                            address_validated=False,
                            status='new'
                        )

                        db.session.add(job)
                        db.session.flush()

                        logger.info(f"✅ Job #{job.id} created: {extracted.get('service_type', 'general')} - {extracted.get('customer_name', 'Unknown')} - booking_needed={booking_needed}")

                        # Step 5: Queue SMS confirmation if phone number available
                        if extracted.get('customer_phone'):
                            sms_message = f"Hi {extracted.get('customer_name', 'there')}! Your service request has been received. Reference: #{job.id}. We'll confirm your appointment shortly."

                            sms_log = SMSLog(
                                job_id=job.id,
                                recipient_phone=extracted.get('customer_phone'),
                                message_type='confirmation',
                                message_body=sms_message,
                                status='queued'
                            )
                            db.session.add(sms_log)
                            logger.info(f"📱 SMS confirmation queued for {extracted.get('customer_phone')}")
                        else:
                            logger.info(f"ℹ️  No phone number - SMS skipped, job created for follow-up")

                        db.session.commit()

                except Exception as e:
                    logger.error(f"❌ Error in CallTradie job creation: {e}")
                    import traceback
                    logger.error(traceback.format_exc())

            # Trigger n8n workflow if configured (NON-BLOCKING)
            if call_log.agent_id and call_log.status == 'completed':
                agent = db.session.get(Agent, call_log.agent_id)

                if agent and agent.workflow_id:
                    workflow = db.session.get(Workflow, agent.workflow_id)

                    if workflow and workflow.is_active:
                        # Prepare comprehensive data for n8n
                        n8n_data = {
                            'event_type': 'call_completed',
                            'call_id': call_log.id,
                            'room_name': room_name,
                            'agent': {
                                'id': agent.id,
                                'name': agent.name
                            },
                            'call_details': {
                                'from_number': call_log.from_number,
                                'to_number': call_log.to_number,
                                'duration_seconds': duration,
                                'duration_minutes': call_log.minutes_used,
                                'status': call_log.status,
                                'call_type': call_log.call_type,
                                'timestamp': call_log.created_at.isoformat()
                            },
                            'conversation': {
                                'transcription': transcription,
                                'message_count': message_count,
                                'recording_url': call_log.recording_url
                            }
                        }

                        # Add sentiment/interest analysis if available
                        if call_log.sentiment_summary:
                            try:
                                sentiment = json.loads(call_log.sentiment_summary)
                                n8n_data['analysis'] = {
                                    'interest_level': sentiment.get('interest'),
                                    'confidence': sentiment.get('confidence'),
                                    'key_indicators': sentiment.get('key_indicators', {})
                                }
                            except:
                                pass

                        # Add campaign info if applicable
                        if is_campaign_call:
                            n8n_data['campaign'] = {
                                'campaign_id': campaign_id,
                                'contact_id': contact_id
                            }

                        # Update workflow stats (will be updated in background thread on success)
                        workflow.total_calls += 1
                        db.session.commit()

                        # Trigger POST-call webhook if enabled (NON-BLOCKING - returns immediately)
                        if workflow.post_call_enabled:
                            logger.info(f"Triggering workflow: {workflow.name}")
                            webhook_service.trigger_webhook(
                                workflow_id=workflow.id,
                                workflow_url=workflow.webhook_url,
                                api_key=workflow.api_key,
                                call_data=n8n_data,
                                call_log_id=call_log.id
                            )
                        else:
                            logger.info(f"Post-call webhook DISABLED for workflow: {workflow.name}")

            return jsonify({
                'success': True,
                'message': f'Call processed: {message_count} messages',
                'transcription_length': len(transcription)
            }), 200

        logger.warning(f"Call log not found: {room_name}")
        return jsonify({'error': 'Call log not found'}), 404

    except Exception as e:
        logger.error(f"Webhook error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

# ==================== HELPER FUNCTIONS ====================

def analyze_sentiment(messages, agent_prompt=None):
    """
    Analyze the entire conversation to determine genuine user interest.
    This analyzes the full conversation context including what the agent was offering.

    Args:
        messages: List of conversation messages
        agent_prompt: The agent's instructions/prompt to understand what was being offered

    Returns: {
        'interest': 'Interested' or 'Not Interested',
        'confidence': 'High' or 'Medium' or 'Low'
    }
    """
    if not messages or len(messages) == 0:
        return {
            'interest': 'Not Interested',
            'confidence': 'Low',
            'reason': 'No conversation data',
            'key_indicators': {
                'asked_questions': False,
                'requested_callback': False,
                'shared_needs': False,
                'agreed_next_steps': False,
                'provided_details': False,
                'explicit_rejection': False
            }
        }

    try:
        from openai import OpenAI
        client = OpenAI(api_key=os.environ.get('OPENAI_API_KEY'))

        # Prepare full conversation for comprehensive analysis
        conversation_text = "\n".join([
            f"{msg['role'].capitalize()}: {msg['text']}"
            for msg in messages
        ])

        # Count messages for context
        user_messages = [msg for msg in messages if msg['role'] == 'user']
        agent_messages = [msg for msg in messages if msg['role'] == 'agent']

        # Include agent context if available
        agent_context = ""
        if agent_prompt:
            agent_context = f"""
AGENT'S INSTRUCTIONS/PURPOSE:
{agent_prompt}

This context helps you understand what the agent was offering/discussing.
"""

        prompt = f"""Analyze the ENTIRE conversation to determine if the user shows GENUINE INTEREST in what was being offered.

{agent_context}

DO NOT judge based solely on the customer's tone or politeness. Analyze the CONTENT and CONTEXT of the full conversation.

Full Conversation ({len(user_messages)} user messages, {len(agent_messages)} agent messages):
{conversation_text}

CRITICAL ANALYSIS CRITERIA - Evaluate the WHOLE conversation:

1. ENGAGEMENT LEVEL:
   - Does the user provide substantive responses? (not just "yes", "no", "ok")
   - Does the user answer the agent's questions with relevant details?
   - Is there meaningful back-and-forth dialogue?

2. INTEREST INDICATORS:
   - Does the user ask questions about the product/service/topic?
   - Does the user request more information, pricing, demos, or callbacks?
   - Does the user share their needs, problems, or use cases?
   - Does the user agree to next steps (meeting, call back, email, etc.)?
   - Does the user show curiosity or want to learn more?

3. CONVERSATION FLOW:
   - Is the user actively participating in the conversation?
   - Does the user contribute meaningfully to the discussion?
   - Does the conversation progress naturally with both parties engaged?

4. CLEAR DISINTEREST SIGNALS (Mark as NOT INTERESTED only if these are present):
   - Explicit rejection: "not interested", "no thanks", "don't call again"
   - Trying to end conversation quickly: "I'm busy", "gotta go", "not now", "wrong number"
   - Avoiding or ignoring questions repeatedly
   - Very brief dismissive responses throughout entire conversation (Make sure sometimes user give short answers but still show interest by asking questions or sharing needs)
   - Asking to be removed from list

DECISION CRITERIA:
 INTERESTED = User shows engagement, asks questions, shares needs, wants follow-up, provides detailed responses, agrees to next steps
L NOT INTERESTED = Clear rejection, trying to end call quickly, dismissive throughout, explicit disinterest

IMPORTANT: Being polite but brief ` NOT INTERESTED
- If user is polite and answers questions, even if brief � lean toward INTERESTED
- If user provides any details about their situation � INTERESTED
- If user doesn't explicitly reject � analyze engagement level carefully
- Focus on CONTENT of what they say, not tone

Return ONLY valid JSON:
{{
    "interest": "Interested" or "Not Interested",
    "confidence": "High" or "Medium" or "Low",
    "reason": "Brief explanation of the decision based on conversation content",
    "key_indicators": {{
        "asked_questions": true/false,
        "requested_callback": true/false,
        "shared_needs": true/false,
        "agreed_next_steps": true/false,
        "provided_details": true/false,
        "explicit_rejection": true/false
    }}
}}"""

        response = client.chat.completions.create(
            model="gpt-4.1-nano",
            messages=[
                {"role": "system", "content": "You are an expert conversation analyst. Analyze customer interest accurately based on conversation content, not tone."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_completion_tokens=500
        )

        result_text = response.choices[0].message.content.strip()

        # Parse JSON response
        import re
        json_match = re.search(r'\{.*\}', result_text, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group())
            logger.info(f"=� Sentiment Analysis: {result.get('interest')} ({result.get('confidence')} confidence)")
            return result
        else:
            logger.warning("� Could not parse sentiment analysis JSON")
            return {
                'interest': 'Not Interested',
                'confidence': 'Low',
                'reason': 'Failed to parse analysis',
                'key_indicators': {
                    'asked_questions': False,
                    'requested_callback': False,
                    'shared_needs': False,
                    'agreed_next_steps': False,
                    'provided_details': False,
                    'explicit_rejection': False
                }
            }

    except Exception as e:
        logger.error(f"Error analyzing sentiment: {e}")
        return {
            'interest': 'Not Interested',
            'confidence': 'Low',
            'reason': f'Analysis error: {str(e)}',
            'key_indicators': {
                'asked_questions': False,
                'requested_callback': False,
                'shared_needs': False,
                'agreed_next_steps': False,
                'provided_details': False,
                'explicit_rejection': False
            }
        }


# ==================== RECORDING DOWNLOAD PROXY ====================

@core_bp.route('/download-recording/<int:call_id>')
@login_required
@approved_required
def download_recording(call_id):
    """Proxy download for call recordings to hide the actual OSS URL"""
    import requests
    from flask import Response, stream_with_context

    # Get the call log
    call_log = CallLog.query.get_or_404(call_id)

    # Check ownership
    if call_log.user_id != session['user_id'] and not session.get('is_admin'):
        flash('Access denied', 'danger')
        return redirect(url_for('core.call_logs'))

    # Check if recording exists
    if not call_log.recording_url:
        flash('No recording available for this call', 'error')
        return redirect(url_for('core.view_call_log', log_id=call_id))

    try:
        # Stream the file from OSS
        r = requests.get(call_log.recording_url, stream=True)
        r.raise_for_status()

        # Create a response with the file stream
        return Response(
            stream_with_context(r.iter_content(chunk_size=8192)),
            content_type='audio/mp4',
            headers={
                'Content-Disposition': f'attachment; filename=recording_{call_id}.mp4',
                'Content-Length': r.headers.get('Content-Length', '')
            }
        )
    except Exception as e:
        logger.error(f"Error downloading recording: {e}")
        flash('Failed to download recording', 'error')
        return redirect(url_for('core.view_call_log', log_id=call_id))